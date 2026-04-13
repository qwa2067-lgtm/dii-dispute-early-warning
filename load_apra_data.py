"""
load_apra_data.py
-----------------
Reads the two APRA Life Insurance Claims and Disputes Statistics spreadsheets
and extracts the data needed for the DII Early Warning Dashboard.

Source:
    APRA Life Insurance Claims and Disputes Statistics (public, Creative Commons CC BY 3.0 AU)
    https://www.apra.gov.au/life-insurance-claims-and-disputes-statistics

Outputs (written to data/ subfolder):
    dispute_trend.csv     — DII dispute lodgement ratio, Ind Non-Advised, 2018–2025
    claims_trend.csv      — DII claims (received, admitted, declined), Ind Non-Advised, 2018–2025
    dispute_by_product.csv — Dispute rates by product, June 2025 (all channels)
    dispute_outcomes.csv  — Dispute resolution outcomes by product, June 2025
    dispute_duration.csv  — Dispute duration by product, June 2025

Usage:
    python3 load_apra_data.py

Requirements:
    pip install pandas openpyxl
"""

import pandas as pd
from pathlib import Path

# ── Config ─────────────────────────────────────────────────────────────────────

FOLDER       = Path(__file__).parent
DATA_FOLDER  = FOLDER / "data"
DATA_FOLDER.mkdir(exist_ok=True)

HISTORICAL_FILE = FOLDER / "Life insurance claims and disputes statistics database June 2018 to June 2025_0.xlsx"
SNAPSHOT_FILE   = FOLDER / "Life insurance claims and disputes data June 2025.xlsx"

CHANNEL     = "Individual Non-Advised"
COVER       = "DII"

# ── Helpers ────────────────────────────────────────────────────────────────────

def read_sheet(path: Path, sheet: str) -> pd.DataFrame:
    """Read a sheet from an APRA workbook into a clean DataFrame."""
    df = pd.read_excel(path, sheet_name=sheet, header=0)
    # Drop fully empty rows and columns
    df = df.dropna(how="all").reset_index(drop=True)
    return df


def read_database_sheet(sheet: str) -> pd.DataFrame:
    """Read a tidy-format sheet from the historical database."""
    wb = pd.ExcelFile(HISTORICAL_FILE)
    raw = wb.parse(sheet, header=0)
    # The database sheets have a header row followed by data
    # Column names are in the first non-null row
    raw.columns = raw.iloc[0] if raw.iloc[0].notna().all() else raw.columns
    raw = raw.iloc[1:].reset_index(drop=True)
    raw['Value'] = pd.to_numeric(raw['Value'], errors='coerce')
    raw['Reporting Date'] = pd.to_datetime(raw['Reporting Date'], errors='coerce')
    raw = raw[raw['Reporting Date'].notna()]
    return raw


# ── Extract: dispute trend (historical database) ───────────────────────────────

def extract_dispute_trend() -> pd.DataFrame:
    """
    DII dispute lodgement ratio (per 100,000 lives insured) over time,
    for Individual Non-Advised channel.
    Aggregated across all insurers (industry total).
    """
    import openpyxl
    wb = openpyxl.load_workbook(HISTORICAL_FILE)
    ws = wb['Disputes']
    data = list(ws.iter_rows(values_only=True))
    df = pd.DataFrame(data[1:], columns=data[0])
    df = df[df['Reporting Date'].notna()].copy()
    df['Value'] = pd.to_numeric(df['Value'], errors='coerce')
    df['Reporting Date'] = pd.to_datetime(df['Reporting Date'], errors='coerce')

    dii = df[
        (df['Cover Type'] == COVER) &
        (df['Channel Type'] == CHANNEL)
    ].copy()

    # Dispute lodgement ratio — average across insurers (it's already a rate per 100k lives)
    ratio = dii[dii['Data item'] == 'Number of disputes per 100,000 lives insured']
    ratio_trend = (
        ratio.groupby('Reporting Date')['Value']
        .mean()
        .reset_index()
        .rename(columns={'Value': 'dispute_rate_per_100k'})
    )

    # Total disputes lodged — sum across insurers
    lodged = dii[
        (dii['Data item'] == 'Number of disputes') &
        (dii['Category'] == 'Disputes Lodged')
    ]
    lodged_trend = (
        lodged.groupby('Reporting Date')['Value']
        .sum()
        .reset_index()
        .rename(columns={'Value': 'disputes_lodged'})
    )

    # Original decision reversed — sum across insurers
    reversed_ = dii[
        (dii['Data item'] == 'Number of disputes') &
        (dii['Category'] == 'Original decision reversed')
    ]
    reversed_trend = (
        reversed_.groupby('Reporting Date')['Value']
        .sum()
        .reset_index()
        .rename(columns={'Value': 'decisions_reversed'})
    )

    # Disputes resolved — sum across insurers
    resolved = dii[
        (dii['Data item'] == 'Number of disputes') &
        (dii['Category'] == 'Disputes Resolved')
    ]
    resolved_trend = (
        resolved.groupby('Reporting Date')['Value']
        .sum()
        .reset_index()
        .rename(columns={'Value': 'disputes_resolved'})
    )

    # Merge all
    trend = (
        ratio_trend
        .merge(lodged_trend, on='Reporting Date', how='left')
        .merge(resolved_trend, on='Reporting Date', how='left')
        .merge(reversed_trend, on='Reporting Date', how='left')
    )
    trend['reversal_rate'] = trend['decisions_reversed'] / trend['disputes_resolved']
    trend['year_label'] = trend['Reporting Date'].dt.strftime('%b %Y')
    trend = trend.sort_values('Reporting Date').reset_index(drop=True)

    return trend


# ── Extract: claims trend (historical database) ────────────────────────────────

def extract_claims_trend() -> pd.DataFrame:
    """
    DII claims statistics over time for Individual Non-Advised channel.
    Includes: received, admitted, declined, withdrawn.
    """
    import openpyxl
    wb = openpyxl.load_workbook(HISTORICAL_FILE)
    ws = wb['Claims']
    data = list(ws.iter_rows(values_only=True))
    df = pd.DataFrame(data[1:], columns=data[0])
    df = df[df['Reporting Date'].notna()].copy()
    df['Value'] = pd.to_numeric(df['Value'], errors='coerce')
    df['Reporting Date'] = pd.to_datetime(df['Reporting Date'], errors='coerce')

    dii = df[
        (df['Cover Type'] == COVER) &
        (df['Channel Type'] == CHANNEL) &
        (df['Data item'] == 'Number of claims')
    ].copy()

    categories = {
        'Total Claims Received':       'claims_received',
        'Finalised Claims - Admitted': 'claims_admitted',
        'Finalised Claims - Declined': 'claims_declined',
        'Withdrawn Claims':            'claims_withdrawn',
    }

    frames = []
    for cat, col in categories.items():
        sub = dii[dii['Category'] == cat].groupby('Reporting Date')['Value'].sum().reset_index()
        sub = sub.rename(columns={'Value': col})
        frames.append(sub)

    trend = frames[0]
    for f in frames[1:]:
        trend = trend.merge(f, on='Reporting Date', how='left')

    trend['admittance_rate'] = trend['claims_admitted'] / trend['claims_received']
    trend['decline_rate']    = trend['claims_declined'] / trend['claims_received']
    trend['year_label']      = trend['Reporting Date'].dt.strftime('%b %Y')
    trend = trend.sort_values('Reporting Date').reset_index(drop=True)

    return trend


# ── Extract: dispute rates by product — June 2025 snapshot ────────────────────

def extract_dispute_by_product() -> pd.DataFrame:
    """
    Dispute lodgement ratio by product for Individual Non-Advised channel,
    as at June 2025. Used to show DII is the most problematic product.
    """
    import openpyxl
    wb = openpyxl.load_workbook(SNAPSHOT_FILE)
    ws = wb['Industry_Level_Results']

    rows = list(ws.iter_rows(values_only=True))

    # Find the dispute lodgement ratio section
    products = ['Death', 'TPD', 'Trauma', 'DII', 'CCI', 'Funeral', 'Accident']
    result = []

    in_section = False
    for row in rows:
        if row[0] and 'Dispute lodgement ratio' in str(row[0]):
            in_section = True
            continue
        if in_section and row[0] in products:
            # Columns: product, Ind Advised, Ind Non-Advised, Group Super, Group Ordinary
            try:
                ind_advised     = float(row[1]) if row[1] not in (None, 'n/a', '*') else None
                ind_non_advised = float(row[2]) if row[2] not in (None, 'n/a', '*') else None
                grp_super       = float(row[3]) if row[3] not in (None, 'n/a', '*') else None
                grp_ordinary    = float(row[4]) if row[4] not in (None, 'n/a', '*') else None
            except (TypeError, ValueError):
                ind_advised = ind_non_advised = grp_super = grp_ordinary = None
            result.append({
                'product':          row[0],
                'ind_advised':      ind_advised,
                'ind_non_advised':  ind_non_advised,
                'grp_super':        grp_super,
                'grp_ordinary':     grp_ordinary,
            })
        if in_section and row[0] and row[0] not in products and row[0] not in (None,) and len(result) > 0:
            # We've passed the section
            break

    return pd.DataFrame(result)


# ── Extract: dispute outcomes — June 2025 snapshot ────────────────────────────

def extract_dispute_outcomes() -> pd.DataFrame:
    """
    Dispute resolution outcomes by product as at June 2025.
    Includes: % resolved, % reversed, % maintained, % other.
    """
    import openpyxl
    wb = openpyxl.load_workbook(SNAPSHOT_FILE)
    ws = wb['Industry_Level_Results']
    rows = list(ws.iter_rows(values_only=True))

    products = ['Death', 'TPD', 'Trauma', 'DII', 'CCI', 'Funeral', 'Accident']
    result = []

    in_section = False
    for row in rows:
        if row[0] and 'Disputes outcomes by cover type' in str(row[0]):
            in_section = True
            continue
        if in_section and row[0] in products:
            try:
                result.append({
                    'product':             row[0],
                    'pct_resolved':        float(row[1]) if row[1] not in (None, 'n/a') else None,
                    'pct_maintained':      float(row[2]) if row[2] not in (None, 'n/a') else None,
                    'pct_reversed':        float(row[3]) if row[3] not in (None, 'n/a') else None,
                    'pct_other_outcome':   float(row[4]) if row[4] not in (None, 'n/a') else None,
                    'pct_withdrawn':       float(row[5]) if row[5] not in (None, 'n/a') else None,
                })
            except (TypeError, ValueError):
                pass
        if in_section and row[0] and row[0] not in products and row[0] not in (None,) and len(result) > 0:
            break

    return pd.DataFrame(result)


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("APRA Data Loader — DII Early Warning Dashboard")
    print("=" * 60)

    print("\n1. Extracting dispute trend (2018–2025)...")
    dispute_trend = extract_dispute_trend()
    dispute_trend.to_csv(DATA_FOLDER / "dispute_trend.csv", index=False)
    print(f"   {len(dispute_trend)} periods. Dispute rate range: "
          f"{dispute_trend['dispute_rate_per_100k'].min():.0f}–"
          f"{dispute_trend['dispute_rate_per_100k'].max():.0f} per 100k lives")
    print(dispute_trend[['year_label', 'dispute_rate_per_100k', 'disputes_lodged', 'reversal_rate']].to_string(index=False))

    print("\n2. Extracting claims trend (2018–2025)...")
    claims_trend = extract_claims_trend()
    claims_trend.to_csv(DATA_FOLDER / "claims_trend.csv", index=False)
    print(f"   {len(claims_trend)} periods. Admittance rate range: "
          f"{claims_trend['admittance_rate'].min():.1%}–"
          f"{claims_trend['admittance_rate'].max():.1%}")

    print("\n3. Extracting dispute rates by product (June 2025)...")
    by_product = extract_dispute_by_product()
    by_product.to_csv(DATA_FOLDER / "dispute_by_product.csv", index=False)
    print(by_product.to_string(index=False))

    print("\n4. Extracting dispute outcomes (June 2025)...")
    outcomes = extract_dispute_outcomes()
    outcomes.to_csv(DATA_FOLDER / "dispute_outcomes.csv", index=False)
    print(outcomes.to_string(index=False))

    print(f"\nAll files saved to {DATA_FOLDER}/")
    print("Next step: run generate_complaints.py")


if __name__ == "__main__":
    main()
